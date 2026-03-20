import openpyxl


fastmovin = openpyxl.load_workbook('.\\Misc\\Fast Moving Data.xlsx')
sheetfm = 'Graphic'
fastbook = fastmovin[sheetfm]

data = openpyxl.load_workbook('.\\Misc\\data.xlsx')
sheetdata = 'fastgpu'
databook = data[sheetdata]

start_row = 2
counter = 0
for row in range(2, fastbook.max_row+1):
	sku = fastbook[f'A{row}'].value
	name = fastbook[f'B{row}'].value
	databook[f'A{start_row + counter * 7}'] = name
	databook[f'E{start_row + counter * 7}'] = sku
	counter+=1

data.save('data.xlsx')