from datetime import datetime
import openpyxl
import sqlite3

wb = openpyxl.load_workbook('.\\Misc\\data.xlsx')
### dont forget to change sheet name for specific categories ###
sheet = 'fastgpu'
ws = wb[sheet]

# Connect to the SQLite database
conn = sqlite3.connect('server/newScraper.db')
cursor = conn.cursor()

# Insert data into the `products` table
for row in range(2, ws.max_row + 1, 7):  # Start at row 2, increment by 7
    product_name = ws[f'A{row}'].value
    product_sku = ws[f'E{row}'].value
    prime_link = ws[f'C{row}'].value
    mdcomp_link = ws[f'C{row + 1}'].value
    vedant_link = ws[f'C{row + 2}'].value
    pcstudio_link = ws[f'C{row + 3}'].value
    clarion_link = ws[f'C{row + 4}'].value
    ehubs_link = ws[f'C{row + 5}'].value

    if prime_link in ['not found', '', None]:
        prime_link = 'product link not found'

    if mdcomp_link in ['not found', '', None]:
        mdcomp_link = 'product link not found'

    if vedant_link in ['not found', '', None]:
        vedant_link = 'product link not found'

    if pcstudio_link in ['not found', '', None]:
        pcstudio_link = 'product link not found'

    if clarion_link in ['not found', '', None]:
        clarion_link = 'product link not found'

    if ehubs_link in ['not found', '', None]:
        ehubs_link = 'product link not found'

# Start at row 2, increment by 8 
# only for FASTMOVING SSD WITH ONLYSSD ADDED IN THE LIST
# for row in range(2, ws.max_row + 1, 8):  
#     product_name = ws[f'A{row}'].value
#     product_sku = ws[f'E{row}'].value
#     prime_link = ws[f'C{row}'].value
#     mdcomp_link = ws[f'C{row + 1}'].value
#     vedant_link = ws[f'C{row + 2}'].value
#     pcstudio_link = ws[f'C{row + 3}'].value
#     clarion_link = ws[f'C{row + 4}'].value
#     ehubs_link = ws[f'C{row + 5}'].value
#     ossd_link = ws[f'C{row + 6}'].value

#     if prime_link in ['not found', '', None]:
#         prime_link = 'product link not found'

#     if mdcomp_link in ['not found', '', None]:
#         mdcomp_link = 'product link not found'

#     if vedant_link in ['not found', '', None]:
#         vedant_link = 'product link not found'

#     if pcstudio_link in ['not found', '', None]:
#         pcstudio_link = 'product link not found'

#     if clarion_link in ['not found', '', None]:
#         clarion_link = 'product link not found'

#     if ehubs_link in ['not found', '', None]:
#         ehubs_link = 'product link not found'

    # Insert data into the `products` table
# SSD
    # cursor.execute('''
    # INSERT OR IGNORE INTO ssd_products (product_name, product_sku, prime_price, mdcomp_price, vedant_price, pcstudio_price, clarion_price, ehubs_price, prime_stock, mdcomp_stock, vedant_stock, pcstudio_stock, clarion_stock, ehubs_stock)
    # VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    # ''', (product_name, product_sku, None, None, None, None, None, None, None, None, None, None, None, None))

    # cursor.execute('''
    # INSERT OR IGNORE INTO ssd_product_links (product_sku, prime_link, mdcomp_link, vedant_link, pcstudio_link, clarion_link, ehubs_link, date_added)
    # VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    # ''', (product_sku, prime_link, mdcomp_link, vedant_link, pcstudio_link, clarion_link, ehubs_link, datetime.now().strftime('%Y-%m-%d')))

# CPU
    # cursor.execute('''
    # INSERT OR IGNORE INTO cpu_products (product_name, product_sku, prime_price, mdcomp_price, vedant_price, pcstudio_price, clarion_price, ehubs_price, prime_stock, mdcomp_stock, vedant_stock, pcstudio_stock, clarion_stock, ehubs_stock)
    # VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    # ''', (product_name, product_sku, None, None, None, None, None, None, None, None, None, None, None, None))

    # cursor.execute('''
    # INSERT OR IGNORE INTO cpu_product_links (product_sku, prime_link, mdcomp_link, vedant_link, pcstudio_link, clarion_link, ehubs_link)
    # VALUES (?, ?, ?, ?, ?, ?, ?)
    # ''', (product_sku, prime_link, mdcomp_link, vedant_link, pcstudio_link, clarion_link, ehubs_link))
    
# RAM
    # cursor.execute('''
    # INSERT OR IGNORE INTO ram_products (product_name, product_sku, prime_price, mdcomp_price, vedant_price, pcstudio_price, clarion_price, ehubs_price, prime_stock, mdcomp_stock, vedant_stock, pcstudio_stock, clarion_stock, ehubs_stock)
    # VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    # ''', (product_name, product_sku, None, None, None, None, None, None, None, None, None, None, None, None))

    # cursor.execute('''
    # INSERT OR IGNORE INTO ram_product_links (product_sku, prime_link, mdcomp_link, vedant_link, pcstudio_link, clarion_link, ehubs_link)
    # VALUES (?, ?, ?, ?, ?, ?, ?)
    # ''', (product_sku, prime_link, mdcomp_link, vedant_link, pcstudio_link, clarion_link, ehubs_link))
 
# GPU
    # cursor.execute('''
    # INSERT OR IGNORE INTO gpu_products (product_name, product_sku, prime_price, mdcomp_price, vedant_price, pcstudio_price, clarion_price, ehubs_price, prime_stock, mdcomp_stock, vedant_stock, pcstudio_stock, clarion_stock, ehubs_stock)
    # VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    # ''', (product_name, product_sku, None, None, None, None, None, None, None, None, None, None, None, None))

    # cursor.execute('''
    # INSERT OR IGNORE INTO gpu_product_links (product_sku, prime_link, mdcomp_link, vedant_link, pcstudio_link, clarion_link, ehubs_link)
    # VALUES (?, ?, ?, ?, ?, ?, ?)
    # ''', (product_sku, prime_link, mdcomp_link, vedant_link, pcstudio_link, clarion_link, ehubs_link))

# CABINET    
    # cursor.execute('''
    # INSERT OR IGNORE INTO cabinet_products (product_name, product_sku, prime_price, mdcomp_price, vedant_price, pcstudio_price, clarion_price, ehubs_price, prime_stock, mdcomp_stock, vedant_stock, pcstudio_stock, clarion_stock, ehubs_stock)
    # VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    # ''', (product_name, product_sku, None, None, None, None, None, None, None, None, None, None, None, None))

    # cursor.execute('''
    # INSERT OR IGNORE INTO cabinet_product_links (product_sku, prime_link, mdcomp_link, vedant_link, pcstudio_link, clarion_link, ehubs_link)
    # VALUES (?, ?, ?, ?, ?, ?, ?)
    # ''', (product_sku, prime_link, mdcomp_link, vedant_link, pcstudio_link, clarion_link, ehubs_link))


################ FAST MOVING #########################
    # FAST MOVING SSD ##
    # cursor.execute('''
    # INSERT OR IGNORE INTO fastmoving_ssd_products (product_name, product_sku, prime_price, ossd_price, mdcomp_price, vedant_price, pcstudio_price, clarion_price, ehubs_price, prime_stock, ossd_stock, mdcomp_stock, vedant_stock, pcstudio_stock, clarion_stock, ehubs_stock)
    # VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    # ''', (product_name, product_sku, None, None, None, None, None, None, None, None, None, None, None, None, None, None))

    # cursor.execute('''
    # INSERT OR IGNORE INTO fastmoving_ssd_product_links (product_sku, prime_link, ossd_link, mdcomp_link, vedant_link, pcstudio_link, clarion_link, ehubs_link, date_added)
    # VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    # ''', (product_sku, prime_link, ossd_link, mdcomp_link, vedant_link, pcstudio_link, clarion_link, ehubs_link, datetime.now().strftime('%Y-%m-%d')))

    # FAST MOVING CPU ##
    # cursor.execute('''
    # INSERT OR IGNORE INTO fastmoving_cpu_products (product_name, product_sku, prime_price, mdcomp_price, vedant_price, pcstudio_price, clarion_price, ehubs_price, prime_stock, mdcomp_stock, vedant_stock, pcstudio_stock, clarion_stock, ehubs_stock)
    # VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    # ''', (product_name, product_sku, None, None, None, None, None, None, None, None, None, None, None, None))

    # cursor.execute('''
    # INSERT OR IGNORE INTO fastmoving_cpu_product_links (product_sku, prime_link, mdcomp_link, vedant_link, pcstudio_link, clarion_link, ehubs_link, date_added)
    # VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    # ''', (product_sku, prime_link, mdcomp_link, vedant_link, pcstudio_link, clarion_link, ehubs_link, datetime.now().strftime('%Y-%m-%d')))

    # FAST MOVING RAM ##
    # cursor.execute('''
    # INSERT OR IGNORE INTO fastmoving_ram_products (product_name, product_sku, prime_price, mdcomp_price, vedant_price, pcstudio_price, clarion_price, ehubs_price, prime_stock, mdcomp_stock, vedant_stock, pcstudio_stock, clarion_stock, ehubs_stock)
    # VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    # ''', (product_name, product_sku, None, None, None, None, None, None, None, None, None, None, None, None))

    # cursor.execute('''
    # INSERT OR IGNORE INTO fastmoving_ram_product_links (product_sku, prime_link, mdcomp_link, vedant_link, pcstudio_link, clarion_link, ehubs_link, date_added)
    # VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    # ''', (product_sku, prime_link, mdcomp_link, vedant_link, pcstudio_link, clarion_link, ehubs_link, datetime.now().strftime('%Y-%m-%d')))

    # FAST MOVING GPU ##
    cursor.execute('''
    INSERT OR IGNORE INTO fastmoving_gpu_products (product_name, product_sku, prime_price, mdcomp_price, vedant_price, pcstudio_price, clarion_price, ehubs_price, prime_stock, mdcomp_stock, vedant_stock, pcstudio_stock, clarion_stock, ehubs_stock)
    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (product_name, product_sku, None, None, None, None, None, None, None, None, None, None, None, None))

    cursor.execute('''
    INSERT OR IGNORE INTO fastmoving_gpu_product_links (product_sku, prime_link, mdcomp_link, vedant_link, pcstudio_link, clarion_link, ehubs_link, date_added)
    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    ''', (product_sku, prime_link, mdcomp_link, vedant_link, pcstudio_link, clarion_link, ehubs_link, datetime.now().strftime('%Y-%m-%d')))



# Commit and close the connection
conn.commit()
conn.close()