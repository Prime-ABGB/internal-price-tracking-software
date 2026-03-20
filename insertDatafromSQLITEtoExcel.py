import sqlite3
from openpyxl import load_workbook

# Connect to SQLite database
# conn = sqlite3.connect('scraper.db')
conn = sqlite3.connect('server/newscraper.db')
cursor = conn.cursor()

# Fetch data from SQLite table where cpuid is between 160 and 292
# SSD
# cursor.execute("SELECT product_name, product_link, sku FROM all_ssds WHERE cpuid BETWEEN 160 AND 438")

# CPU
# cursor.execute("SELECT product_name, product_link, sku FROM all_processors WHERE cpuid BETWEEN 65 AND 182")

# RAM
# cursor.execute("SELECT product_name, product_link, sku FROM all_rams WHERE cpuid BETWEEN 61 AND 368")

# GPU
# cursor.execute("SELECT product_name, product_link, sku FROM all_gpus WHERE cpuid BETWEEN 578 AND 1154")
# NEW filtered data for gpu, now deleted
# cursor.execute("SELECT p.product_sku AS product_sku, p.product_name, pl.prime_link FROM gpu_products p JOIN gpu_product_links pl ON p.product_sku = pl.product_sku WHERE p.prime_stock = 'In Stock'")

rows = cursor.fetchall()

# Close SQLite connection
conn.close()

# Load the existing Excel workbook
wb = load_workbook('.\\Misc\\data.xlsx')

# ssd sheet
# sheet = 'ssd';

# cpu sheet
# sheet = 'cpu'

# ram sheet
# sheet = 'ram'

# gpu sheet
sheet = 'gpu'

ws = wb[sheet]

# Define starting row for data insertion
start_row = 2

# Insert fetched data into Excel starting from the specified row and columns
for index, row_data in enumerate(rows):
  
    ws[f'A{start_row + index * 7}'] = row_data[1]
    
    ws[f'C{start_row + index * 7}'] = row_data[2]
    
    ws[f'E{start_row + index * 7}'] = row_data[0]

# Save the updated Excel file
wb.save('data.xlsx')
